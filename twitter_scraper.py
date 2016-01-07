# -*- coding: utf-8 -*-
import sys
import time
import datetime
from twython import Twython, TwythonError, TwythonRateLimitError
from ignoreconstants import ignore_openpyxl_constants
ignore_openpyxl_constants()
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Style, Font
from urlunshort import resolve
from urlparse import urlsplit
from bs4 import BeautifulSoup
import os
import re
import requests
from requests.exceptions import ConnectionError, MissingSchema, InvalidSchema, ReadTimeout,\
        TooManyRedirects, InvalidURL

CONSUMER_KEY    = "jJc9gVd4GhVpWkoTKRThJmgtg"
CONSUMER_SECRET = "OlNRGpQaXTGj7gIs3FRoxa1hCCoY1VEAMVSThlf9qvzCKkrnSN"
TOKEN_KEY       = "4517074753-t7foTG8zbKyUntKIq6ctam7a2BLoITwPvpR4Ihn"
TOKEN_SECRET    = "UI8czuIJ4l2vHwD9Rb5cdDPq2KZd3LpzMHKQ5MlvdEW45"

USER = 'user'
USERID = 'id_str'
MESSAGE = 'text'
USERURL = 'url'
META = 'search_metadata'
NEXTRESULT = 'next_results'
STATUSES = 'statuses'
EXCELFILE = 'scraped_content.xlsx'
SHEET = 'Scraped Tweets'

#Excel first row Names
ROW_ID = 'User ID'
ROW_USERNAME = 'Username'
ROW_TWEET_MESSAGE = 'Tweet Message'
ROW_KEYPHRASE = 'Keyphrase'
ROW_DATA_RETRIEVED = 'Data Retrieved'
ROW_WEBSITE = 'Website from User'
ROW_MAIL = 'E-Mail from Website'
SHEET_NAME = 'Sheet'
#MAIL_REGEX = re.compile('[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+')
MAIL_REGEX = re.compile('[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$')
LINK_REGEX = re.compile('http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+')

KNOWN_WEBSITES = ['facebook', 'instagram', 'youtube', 'twitter', 'pinterest', 'github', 'tumblr', \
        'yahoo', 'soundcloud', 'amazon', 'apple', 'itunes', 'mtv', 'play.google', 'ask.fm', \
        'soundclick', 'bbc.co', 'news', 'google', 'talkmuzik', 'paypal', 'myspace', 'javascript', \
        'ebay', 'netflix', 'hulu', 'blogspot', 'porn', 'sex', 'hotmovies', 'cigarettes']

def createExcelFile():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=ROW_ID).font = Font(bold=True)
    ws.cell(row=1, column=2, value=ROW_USERNAME).font = Font(bold=True)
    ws.cell(row=1, column=3, value=ROW_TWEET_MESSAGE).font = Font(bold=True)
    ws.cell(row=1, column=4, value=ROW_KEYPHRASE).font = Font(bold=True)
    ws.cell(row=1, column=5, value=ROW_DATA_RETRIEVED).font = Font(bold=True)
    ws.cell(row=1, column=6, value=ROW_WEBSITE).font = Font(bold=True)
    ws.cell(row=1, column=7, value=ROW_MAIL).font = Font(bold=True)
    wb.save(EXCELFILE)

    return wb, ws

def getExcelSheet():
    if os.path.isfile(EXCELFILE):
        print "A %s File already exists, the Scraped content will be appended!" %EXCELFILE
        return load_workbook(EXCELFILE), None
    else:
        return createExcelFile() 

def writeExcelSheet(user_id, username, message, keyword, url, email):
    wb, ws = getExcelSheet()
    if ws is not None:
        ws = writeCells(ws, user_id, username, message, keyword, url, email)
    else:
        ws = wb[SHEET_NAME]
        ws = writeCells(ws, user_id, username, message, keyword, url, email)
    wb.save(EXCELFILE)

def writeCells(ws, user_id, username, message, keyword, url, email):
    sheet_copy = ws
    row_number = ws.max_row+1
    url = resolve(url)
    date_now = datetime.datetime.now()
    formated_time = date_now.strftime('%Y-%m-%d %H:%M')
    sheet_copy.cell(row=row_number, column=1, value=user_id)
    sheet_copy.cell(row=row_number, column=2, value=username.strip().encode('utf-8'))
    sheet_copy.cell(row=row_number, column=3, value=message.strip().encode('utf-8'))
    sheet_copy.cell(row=row_number, column=4, value=keyword)
    sheet_copy.cell(row=row_number, column=5, value=formated_time)
    sheet_copy.cell(row=row_number, column=6, value=url)
    sheet_copy.cell(row=row_number, column=7, value=email)
    return sheet_copy

def getMaxID(response):
    maxId = response[META][NEXTRESULT].split('&')[0].split('?max_id=')[1]
    return maxId

def validateLink(link):
    for h in KNOWN_WEBSITES:
        if h in link and not link.endswith('.jpg'):
            return False
    return True

def extractLinks(response, base_url):
    soup = BeautifulSoup(response.text, 'html.parser')
    links = set([base_url])
    for a in soup.findAll('a', href = LINK_REGEX):
        if 'href' in a.attrs:
            is_valid = validateLink(a.attrs['href'])
            if is_valid:
                links.add(a.attrs['href'])
            else:
                continue
    return links

def checkHostname(hostname):
    for h in KNOWN_WEBSITES:
        if h in hostname:
            return False
    return True

def verifyUrl(url):
    resolved_url = resolve(url)
    if resolved_url is not None:
        parts = urlsplit(resolved_url)
        hostname = parts.hostname
        valid = checkHostname(hostname)
        if valid: 
            base_url = "{0.scheme}://{0.netloc}".format(parts)
            return base_url
        else:
            return None
    else:
        parts = urlsplit(url)
        hostname = parts.hostname
        valid = checkHostname(hostname)
        if valid: 
            base_url = "{0.scheme}://{0.netloc}".format(parts)
            return base_url
        else:
            return None

def requestUrl(url):
    try:
        return requests.get(url, timeout=(5, 30))
    except (InvalidSchema, MissingSchema, ReadTimeout, InvalidURL, ConnectionError) as e:
        print "Website not responding Skipping..."
        print e
        return None

def processUrl(url):
    base_url = verifyUrl(url)
    if base_url is not None:
        response = requestUrl(base_url.strip())
        if response is not None:
            print "Harvesting Links for Domain: %s" %base_url
            return extractLinks(response, base_url)
    else:
        print "Invalid Website Skipping..."

def extractMailfromLinks(links):
    while links:
        link = links.pop()
        print "Searching for e-mail from Link: %s" %link
        try:
            response = requests.get(link.strip(), timeout=(5, 30))
        except (InvalidSchema, MissingSchema, ReadTimeout, ConnectionError, TooManyRedirects) as e:
            print "Invalid Link Skipping..."
            print e
            continue
        soup = BeautifulSoup(response.text, 'html.parser')
        for a in soup.findAll('a', limit=1, href=MAIL_REGEX):
            email = a.attrs['href']
            print "Mail found: %s" %email
            return email
    return None

def getMail(url):
    links = processUrl(url)
    mail = extractMailfromLinks(links)
    if mail is not None:
        return mail
    else:
        return None

def parseTweetStatuses(response, keyword, user_id_list):
    user_id_list = user_id_list 
    for tweet in response[STATUSES]:
        if tweet[USER][USERID] not in user_id_list and tweet[USER][USERURL]:
            user_id_list.append(tweet[USER][USERID])
            scraped_email = getMail(tweet[USER][USERURL])
            if scraped_email is not None:
                writeExcelSheet(tweet[USER][USERID], tweet[USER]['name'], tweet[MESSAGE],
                        keyword, tweet[USER][USERURL], scraped_email)
            else:
                continue
        else:
            continue
    return user_id_list

def mainScraping(tweetobj, keyword, limit=None, user_id_list = [], r_type='recent'):
    #user_id_list = []
    if limit:
        response = tweetobj.search(q=keyword, count=limit)
        parseTweetStatuses(response, keyword)
    else:
        try:
            response = tweetobj.search(q=keyword, result_type=r_type, count=100)
            while NEXTRESULT in response[META].keys():
                maxId = getMaxID(response) 
                response = tweetobj.search(q=keyword, result_type=r_type, max_id=int(maxId)-1, count=100)
                user_id_list = parseTweetStatuses(response, keyword, user_id_list)
        except TwythonRateLimitError as e:
            remaining = float(tweetobj.get_lastfunction_header(header='x-rate-limit-reset')) - time.time()
            print "Twitter Rate Limit Exceeded! Sleeping for %.2f sec." %remaining
            time.sleep(remaining)
            mainScraping(tweetobj, keyword, user_id_list, r_type)
        except TwythonError as e:
            print "Twitter disconnected, reconnecting..."
            reconnectToTwitter(keyword, user_id_list, r_type)
            
def reconnectToTwitter(keyword, user_id_list, r_type):
    tweetobj = Twython(CONSUMER_KEY, CONSUMER_SECRET, TOKEN_KEY, TOKEN_SECRET, client_args={'headers':{'User-Agent': 'MFResearch'}})
    mainScraping(tweetobj, keyword, user_id_list = user_id_list, r_type=r_type)

def main():
    keyphrase = raw_input("Enter Keyword to scrape for: ")
    result_type = raw_input("Select the Result Type [r]ecent, [m]ixed, [p]opular:")
    if result_type in ['r', 'm', 'p']:
        if result_type == 'r':
            result_type = 'recent'
        elif result_type == 'm':
            result_type = 'mixed'
        else:
            result_type = 'popular'

        tweetobj = Twython(CONSUMER_KEY, CONSUMER_SECRET, TOKEN_KEY, TOKEN_SECRET, client_args={'headers':{'User-Agent': 'MFResearch'}})
        mainScraping(tweetobj, keyword=keyphrase)
        print "********************************************"
        print "Scraping finished for keyword %s" %keyphrase
        print "All possible tweets retrieved for this run"
    else:
        print "You Entered a invalid Result Type!!!"
    
if __name__ == "__main__":
    main()
